"""
PharmaSim Excel data parser.

Parses all downloaded XLS reports from PharmaSim into structured Python
dataclasses, organized by year. Each report type has its own parser function
and corresponding dataclass(es).

Usage:
    from parser import load_year
    year0 = load_year(0)
    year1 = load_year(1)

    # Access specific data:
    year0.performance_summary.unit_sales
    year0.income_statement.net_income
    year0.brand_formulations["Allround"].analgesic_mg
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DOWNLOADS_DIR = Path(__file__).parent.parent / "runs"

CHANNELS = [
    "independent_drugstores",
    "chain_drugstores",
    "grocery_stores",
    "convenience_stores",
    "mass_merchandisers",
]

CHANNEL_LABELS = [
    "Indep. Drugstores",
    "Chain Drugstores",
    "Grocery Stores",
    "Convenience Stores",
    "Mass Merchandisers",
]

COMPANIES = ["Allstar", "B & B", "Curall", "Driscol", "Ethik"]

# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _cell(ws, row, col):
    """Read a cell value, returning None for empty cells."""
    v = ws.cell(row=row, column=col).value
    if isinstance(v, str):
        v = v.strip()
    return v


def _float(v):
    """Coerce to float, returning None if not numeric."""
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _open(year: int, name: str, downloads_dir: Path | None = None):
    """Open a workbook by year and report name."""
    base = downloads_dir or DOWNLOADS_DIR
    label = "Year0" if year == 0 else f"Year{year}"
    path = base / f"{label}_{name}.xlsx"
    if not path.exists():
        return None
    return load_workbook(path, data_only=True)


# ===================================================================
# DATACLASSES
# ===================================================================

# -- Performance Summary ------------------------------------------------

@dataclass
class PerformanceSummary:
    """Company-level performance summary for Allround + totals."""
    msrp: float | None = None
    average_discount_pct: float | None = None
    unit_sales: float | None = None
    manufacturer_sales: float | None = None
    promotional_allowance: float | None = None
    cost_of_goods_sold: float | None = None
    gross_margin: float | None = None
    promotional_expenditures: float | None = None
    advertising_expenditures: float | None = None
    product_contribution: float | None = None
    sales_force_expense: float | None = None
    admin_expense: float | None = None
    fixed_costs: float | None = None
    net_income: float | None = None
    cumulative_net_income: float | None = None
    stock_price: float | None = None
    marketing_efficiency_index: float | None = None
    capacity_utilization_pct: float | None = None
    market_share_unit_pct: float | None = None
    market_share_dollar_pct: float | None = None
    average_shelf_space_pct: float | None = None
    # Industry-level data
    industry_total_mfr_sales: float | None = None
    industry_unit_sales: float | None = None
    industry_sales_force_spending: float | None = None
    industry_advertising_spending: float | None = None
    industry_digital_marketing: float | None = None
    industry_promotional_spending: float | None = None
    industry_total_net_income: float | None = None


# -- Income Statement ---------------------------------------------------

@dataclass
class IncomeStatement:
    """Allstar company income statement."""
    manufacturer_sales: float | None = None
    promotional_allowance: float | None = None
    cost_of_goods_sold: float | None = None
    gross_margin: float | None = None
    consumer_and_trade_promo: float | None = None
    advertising: float | None = None
    sales_force: float | None = None
    admin: float | None = None
    total_marketing: float | None = None
    contribution_after_marketing: float | None = None
    fixed_costs: float | None = None
    net_income: float | None = None
    next_year_budget: float | None = None
    # Ratios (% of Mfr Sales)
    promotional_allowance_pct: float | None = None
    cogs_pct: float | None = None
    gross_margin_pct: float | None = None
    consumer_trade_promo_pct: float | None = None
    advertising_pct: float | None = None
    sales_force_pct: float | None = None
    admin_pct: float | None = None
    total_marketing_pct: float | None = None
    contribution_after_marketing_pct: float | None = None
    fixed_costs_pct: float | None = None
    net_income_pct: float | None = None


# -- Product Contribution -----------------------------------------------

@dataclass
class ProductContribution:
    """Product contribution report for a brand (overall and per-unit)."""
    brand: str = ""
    # Overall (in thousands or millions depending on metric)
    unit_sales: float | None = None
    avg_retail_price: float | None = None
    retail_sales: float | None = None
    manufacturer_sales: float | None = None
    allowance_expense: float | None = None
    cost_of_goods_sold: float | None = None
    gross_margin: float | None = None
    consumer_and_trade_promo: float | None = None
    advertising: float | None = None
    total_marketing: float | None = None
    product_contribution: float | None = None
    # Per-unit
    per_unit_avg_retail_price: float | None = None
    per_unit_retail_sales: float | None = None
    per_unit_manufacturer_sales: float | None = None
    per_unit_allowance_expense: float | None = None
    per_unit_cogs: float | None = None
    per_unit_gross_margin: float | None = None
    per_unit_consumer_trade_promo: float | None = None
    per_unit_advertising: float | None = None
    per_unit_total_marketing: float | None = None
    per_unit_product_contribution: float | None = None


# -- Sales Report -------------------------------------------------------

@dataclass
class DiscountTierSales:
    """Sales data for one discount tier."""
    price: float | None = None
    units: float | None = None
    dollars: float | None = None
    pct_of_total: float | None = None


@dataclass
class ChannelSalesDetail:
    """Sales data for one distribution channel."""
    sales_force_count: float | None = None
    units: float | None = None
    dollars: float | None = None
    pct_of_total: float | None = None


@dataclass
class BrandSalesReport:
    """Sales report for a single brand."""
    brand: str = ""
    # By discount schedule
    direct_under_250: DiscountTierSales | None = None
    direct_under_2500: DiscountTierSales | None = None
    direct_2500_plus: DiscountTierSales | None = None
    indirect_wholesale: DiscountTierSales | None = None
    # By distribution channel
    channel_independent_drugstores: ChannelSalesDetail | None = None
    channel_chain_drugstores: ChannelSalesDetail | None = None
    channel_grocery_stores: ChannelSalesDetail | None = None
    channel_convenience_stores: ChannelSalesDetail | None = None
    channel_mass_merchandisers: ChannelSalesDetail | None = None
    channel_wholesalers: ChannelSalesDetail | None = None


# -- Promotion Report ---------------------------------------------------

@dataclass
class BrandPromotionReport:
    """Promotion report for a single brand."""
    brand: str = ""
    promotional_allowance: float | None = None
    coop_advertising: float | None = None
    point_of_purchase: float | None = None
    trial_size: float | None = None
    coupon_expiration_months: float | None = None
    coupon_amount: float | None = None
    trade_rating: float | None = None
    pct_participating_retailers: float | None = None
    coupons_mailed_thousands: float | None = None
    coupon_conversions: float | None = None
    coupons_redeemed: float | None = None
    promo_pct_of_total_sales: float | None = None


# -- Dashboard ----------------------------------------------------------

@dataclass
class Dashboard:
    """Executive dashboard summary."""
    stock_price_previous: float | None = None
    stock_price_current: float | None = None
    stock_price_change: float | None = None
    unit_sales_previous: float | None = None
    unit_sales_current: float | None = None
    unit_sales_change: float | None = None
    revenue_previous: float | None = None
    revenue_current: float | None = None
    revenue_change: float | None = None
    net_income_previous: float | None = None
    net_income_current: float | None = None
    net_income_change: float | None = None
    cumulative_net_income_previous: float | None = None
    cumulative_net_income_current: float | None = None
    cumulative_net_income_change: float | None = None
    market_update_text: str = ""
    company_messages_text: str = ""


# -- Brand Formulations -------------------------------------------------

@dataclass
class BrandFormulation:
    """Ingredient formulation for a single brand."""
    brand: str = ""
    analgesic_mg: float | None = None
    antihistamine_mg: float | None = None
    decongestant_mg: float | None = None
    cough_suppressant_mg: float | None = None
    expectorant_mg: float | None = None
    alcohol_pct: float | None = None
    description: str = ""


# -- Symptoms Reported --------------------------------------------------

@dataclass
class SymptomsReported:
    """Survey of symptoms reported by consumers."""
    aches_pct: float | None = None
    nasal_congestion_pct: float | None = None
    chest_congestion_pct: float | None = None
    runny_nose_pct: float | None = None
    coughing_pct: float | None = None
    allergy_symptoms_pct: float | None = None


# -- Industry Outlook ---------------------------------------------------

@dataclass
class IndustryOutlook:
    """Industry outlook and economic data."""
    population_growth_pct: float | None = None
    industry_growth_pct: float | None = None
    inflation_rate_pct: float | None = None
    population_growth_forecast: str = ""
    industry_growth_forecast: str = ""
    inflation_rate_forecast: str = ""
    pct_direct_sales: float | None = None
    pct_wholesale_sales: float | None = None
    sales_rep_salary: float | None = None
    sales_rep_expenses: float | None = None
    sales_rep_training: float | None = None
    # Marketing research costs
    research_costs: dict[str, float] = field(default_factory=dict)


# -- Advertising --------------------------------------------------------

@dataclass
class BrandAdvertising:
    """Advertising data for a single brand."""
    brand: str = ""
    media_expenditure: float | None = None
    ad_agency: str = ""
    primary_pct: float | None = None
    benefits_pct: float | None = None
    comparison_pct: float | None = None
    reminder_pct: float | None = None


# -- Brand Perceptions --------------------------------------------------

@dataclass
class BrandPerception:
    """Consumer perception scores for a brand (1-5 scale)."""
    brand: str = ""
    aches: float | None = None
    nasal_congestion: float | None = None
    chest_congestion: float | None = None
    runny_nose: float | None = None
    coughing: float | None = None
    allergies: float | None = None


# -- Brand Awareness ----------------------------------------------------

@dataclass
class BrandAwareness:
    """Consumer awareness and trial data for a brand."""
    brand: str = ""
    brand_awareness_pct: float | None = None
    brand_trial_pct: float | None = None
    most_frequent_purchase_pct: float | None = None
    conversion_ratio: float | None = None
    retention_ratio: float | None = None


# -- Shelf Space --------------------------------------------------------

@dataclass
class BrandShelfSpace:
    """Shelf space allocation (%) for a brand by channel."""
    brand: str = ""
    independent_drugstores: float | None = None
    chain_drugstores: float | None = None
    grocery_stores: float | None = None
    convenience_stores: float | None = None
    mass_merchandisers: float | None = None
    all_channels: float | None = None


# -- Promotion (competitor) ---------------------------------------------

@dataclass
class BrandPromotion:
    """Promotion data for a brand (all competitors)."""
    brand: str = ""
    promotional_allowance_pct: float | None = None
    coop_advertising: float | None = None
    point_of_purchase: float | None = None
    trial_size: str = ""
    coupon_amount: float | None = None
    trade_rating: float | None = None


# -- Conjoint Analysis --------------------------------------------------

@dataclass
class ConjointAnalysis:
    """Conjoint analysis results."""
    price_points: list[float] = field(default_factory=list)
    product_rankings: dict[str, list[int]] = field(default_factory=dict)
    product_utilities: dict[str, float] = field(default_factory=dict)
    price_utilities: dict[float, float] = field(default_factory=dict)
    product_importance_pct: float | None = None
    price_importance_pct: float | None = None


# -- Satisfaction -------------------------------------------------------

@dataclass
class BrandSatisfaction:
    """Consumer satisfaction score for a brand."""
    brand: str = ""
    cross_section: float | None = None
    overall: float | None = None


# -- Sales Force --------------------------------------------------------

@dataclass
class CompanySalesForce:
    """Sales force allocation for a company."""
    company: str = ""
    # Direct channels (headcount)
    direct_independent_drugstores: float | None = None
    direct_chain_drugstores: float | None = None
    direct_grocery_stores: float | None = None
    direct_convenience_stores: float | None = None
    direct_mass_merchandisers: float | None = None
    total_direct: float | None = None
    # Indirect
    indirect_wholesaler_support: float | None = None
    indirect_merchandisers: float | None = None
    indirect_detailers: float | None = None
    total_indirect: float | None = None
    total_sales_force: float | None = None
    # Percent allocation (second sheet)
    pct_independent_drugstores: float | None = None
    pct_chain_drugstores: float | None = None
    pct_grocery_stores: float | None = None
    pct_convenience_stores: float | None = None
    pct_mass_merchandisers: float | None = None
    pct_total_direct: float | None = None
    pct_wholesaler_support: float | None = None
    pct_merchandisers: float | None = None
    pct_detailers: float | None = None
    pct_total_indirect: float | None = None


# -- Pricing ------------------------------------------------------------

@dataclass
class BrandPricing:
    """Pricing data for a brand across channels."""
    brand: str = ""
    msrp: float | None = None
    avg_retail_independent_drugstores: float | None = None
    avg_retail_chain_drugstores: float | None = None
    avg_retail_grocery_stores: float | None = None
    avg_retail_convenience_stores: float | None = None
    avg_retail_mass_merchandisers: float | None = None


@dataclass
class ChannelDiscountTier:
    """Units sold at a specific discount tier for a channel."""
    discount_pct: float | None = None
    units_independent_drugstores: float | None = None
    units_chain_drugstores: float | None = None
    units_grocery_stores: float | None = None
    units_convenience_stores: float | None = None
    units_mass_merchandisers: float | None = None


@dataclass
class BrandChannelDiscountDetail:
    """Channel discount detail for a brand."""
    brand: str = ""
    tiers: list[ChannelDiscountTier] = field(default_factory=list)


# -- Decision Criteria --------------------------------------------------

@dataclass
class DecisionCriteria:
    """Consumer decision-making criteria survey."""
    market_penetration_pct: float | None = None
    avg_purchase_per_year: float | None = None
    # Rankings (% who ranked each criterion 1st through 5th)
    effectiveness_rankings: list[float] = field(default_factory=list)
    side_effects_rankings: list[float] = field(default_factory=list)
    price_rankings: list[float] = field(default_factory=list)
    form_rankings: list[float] = field(default_factory=list)
    duration_rankings: list[float] = field(default_factory=list)


# -- Recommendations ----------------------------------------------------

@dataclass
class BrandRecommendation:
    """Physician/pharmacist recommendation rates for a brand."""
    brand: str = ""
    cold_pct: float | None = None
    cough_pct: float | None = None
    allergy_pct: float | None = None


# -- Operating Statistics -----------------------------------------------

@dataclass
class CompanyOperatingStats:
    """Operating statistics for a company."""
    company: str = ""
    retail_sales: float | None = None
    manufacturer_sales: float | None = None
    promotional_allowance: float | None = None
    cost_of_goods_sold: float | None = None
    gross_margin: float | None = None
    consumer_and_trade_promo: float | None = None
    advertising: float | None = None
    sales_force: float | None = None
    admin: float | None = None
    contribution_after_marketing: float | None = None
    fixed_costs: float | None = None
    net_income: float | None = None
    stock_price: float | None = None
    capacity_utilization_pct: float | None = None
    # As percent of retail sales
    pct_retail_manufacturer_sales: float | None = None
    pct_retail_promo_allowance: float | None = None
    pct_retail_cogs: float | None = None
    pct_retail_gross_margin: float | None = None
    pct_retail_consumer_trade_promo: float | None = None
    pct_retail_advertising: float | None = None
    pct_retail_sales_force: float | None = None
    pct_retail_admin: float | None = None
    pct_retail_contrib_after_marketing: float | None = None
    pct_retail_fixed_costs: float | None = None
    pct_retail_net_income: float | None = None
    # As percent of manufacturer sales
    pct_mfr_promo_allowance: float | None = None
    pct_mfr_cogs: float | None = None
    pct_mfr_gross_margin: float | None = None
    pct_mfr_consumer_trade_promo: float | None = None
    pct_mfr_advertising: float | None = None
    pct_mfr_sales_force: float | None = None
    pct_mfr_admin: float | None = None
    pct_mfr_contrib_after_marketing: float | None = None
    pct_mfr_fixed_costs: float | None = None
    pct_mfr_net_income: float | None = None


# -- Manufacturer Sales -------------------------------------------------

@dataclass
class BrandManufacturerSales:
    """Manufacturer sales by market segment for a brand."""
    brand: str = ""
    cold_share_pct: float | None = None
    cough_share_pct: float | None = None
    allergy_share_pct: float | None = None
    nasal_spray_share_pct: float | None = None
    total_share_pct: float | None = None
    cold_sales: float | None = None
    cough_sales: float | None = None
    allergy_sales: float | None = None
    nasal_spray_sales: float | None = None
    total_sales: float | None = None


# -- Channel Sales ------------------------------------------------------

@dataclass
class BrandChannelSales:
    """Retail sales by channel for a brand."""
    brand: str = ""
    # Dollar amounts
    independent_drugstores_sales: float | None = None
    chain_drugstores_sales: float | None = None
    grocery_stores_sales: float | None = None
    convenience_stores_sales: float | None = None
    mass_merchandisers_sales: float | None = None
    total_sales: float | None = None
    # Market share percentages
    independent_drugstores_share_pct: float | None = None
    chain_drugstores_share_pct: float | None = None
    grocery_stores_share_pct: float | None = None
    convenience_stores_share_pct: float | None = None
    mass_merchandisers_share_pct: float | None = None
    total_share_pct: float | None = None


# -- Purchase Intentions ------------------------------------------------

@dataclass
class BrandPurchaseIntention:
    """Purchase intentions vs actual purchases for a brand."""
    brand: str = ""
    intended_pct: float | None = None
    bought_pct: float | None = None


# -- Shopping Habits ----------------------------------------------------

@dataclass
class ShoppingHabits:
    """Consumer shopping habits by channel and segment."""
    # Each is a dict of channel -> percentage
    cold: dict[str, float] = field(default_factory=dict)
    cough: dict[str, float] = field(default_factory=dict)
    allergy: dict[str, float] = field(default_factory=dict)


# -- Trade-Offs ---------------------------------------------------------

@dataclass
class BrandTradeOff:
    """Trade-off survey data for a brand."""
    brand: str = ""
    msrp: float | None = None
    perceived_price: str = ""
    perceived_effectiveness: str = ""
    purchased_pct: float | None = None


# -- Brands Purchased ---------------------------------------------------

@dataclass
class BrandPurchaseSummary:
    """Market-level brand purchase share."""
    brand: str = ""
    cross_section_pct: float | None = None
    overall_pct: float | None = None


@dataclass
class BrandPurchaseDetail:
    """Detailed brand purchase breakdown by segment and demographic."""
    brand: str = ""
    market_units: float | None = None
    brand_units: float | None = None
    # By symptom segment
    cold_market_pct: float | None = None
    cold_brand_share_pct: float | None = None
    cold_brand_pct: float | None = None
    cough_market_pct: float | None = None
    cough_brand_share_pct: float | None = None
    cough_brand_pct: float | None = None
    allergy_market_pct: float | None = None
    allergy_brand_share_pct: float | None = None
    allergy_brand_pct: float | None = None
    # By demographic
    young_singles_market_pct: float | None = None
    young_singles_brand_share_pct: float | None = None
    young_singles_brand_pct: float | None = None
    young_families_market_pct: float | None = None
    young_families_brand_share_pct: float | None = None
    young_families_brand_pct: float | None = None
    mature_families_market_pct: float | None = None
    mature_families_brand_share_pct: float | None = None
    mature_families_brand_pct: float | None = None
    empty_nesters_market_pct: float | None = None
    empty_nesters_brand_share_pct: float | None = None
    empty_nesters_brand_pct: float | None = None
    retired_market_pct: float | None = None
    retired_brand_share_pct: float | None = None
    retired_brand_pct: float | None = None


# -- Year Data Container ------------------------------------------------

@dataclass
class YearData:
    """All parsed data for a single simulation year."""
    year: int = 0
    performance_summary: PerformanceSummary | None = None
    income_statement: IncomeStatement | None = None
    product_contributions: list[ProductContribution] = field(default_factory=list)
    sales_reports: dict[str, BrandSalesReport] = field(default_factory=dict)
    promotion_reports: dict[str, BrandPromotionReport] = field(default_factory=dict)
    dashboard: Dashboard | None = None
    brand_formulations: dict[str, BrandFormulation] = field(default_factory=dict)
    symptoms_reported: SymptomsReported | None = None
    industry_outlook: IndustryOutlook | None = None
    advertising: dict[str, BrandAdvertising] = field(default_factory=dict)
    brand_perceptions: dict[str, BrandPerception] = field(default_factory=dict)
    brand_awareness: dict[str, BrandAwareness] = field(default_factory=dict)
    shelf_space: dict[str, BrandShelfSpace] = field(default_factory=dict)
    promotions: dict[str, BrandPromotion] = field(default_factory=dict)
    conjoint_analysis: ConjointAnalysis | None = None
    satisfaction: dict[str, BrandSatisfaction] = field(default_factory=dict)
    sales_force: dict[str, CompanySalesForce] = field(default_factory=dict)
    pricing: dict[str, BrandPricing] = field(default_factory=dict)
    channel_discount_details: dict[str, BrandChannelDiscountDetail] = field(default_factory=dict)
    decision_criteria: DecisionCriteria | None = None
    recommendations: dict[str, BrandRecommendation] = field(default_factory=dict)
    operating_statistics: dict[str, CompanyOperatingStats] = field(default_factory=dict)
    manufacturer_sales: dict[str, BrandManufacturerSales] = field(default_factory=dict)
    channel_sales: dict[str, BrandChannelSales] = field(default_factory=dict)
    purchase_intentions: dict[str, BrandPurchaseIntention] = field(default_factory=dict)
    shopping_habits: ShoppingHabits | None = None
    trade_offs: dict[str, BrandTradeOff] = field(default_factory=dict)
    brands_purchased_summary: dict[str, BrandPurchaseSummary] = field(default_factory=dict)
    brands_purchased_detail: dict[str, BrandPurchaseDetail] = field(default_factory=dict)


# ===================================================================
# PARSERS
# ===================================================================

def parse_performance_summary(year: int) -> PerformanceSummary | None:
    wb = _open(year, "Performance_Summary")
    if wb is None:
        return None
    ws = wb.active
    ps = PerformanceSummary()

    def _best_val(r):
        """Get best value from cols B or C (some rows use B, some use C)."""
        v = _float(_cell(ws, r, 2))
        if v is None:
            v = _float(_cell(ws, r, 3))
        return v

    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if not label:
            continue
        ll = str(label).lower().strip()
        val = _best_val(r)

        if "suggested retail" in ll or ll == "msrp":
            ps.msrp = val
        elif "average discount" in ll:
            ps.average_discount_pct = val
        elif ll == "unit sales":
            ps.unit_sales = val
        elif ll == "manufacturer sales" or ll == "mfr sales" or ll.startswith("mfr"):
            ps.manufacturer_sales = val
        elif "promo" in ll and "allow" in ll:
            ps.promotional_allowance = val
        elif "cost of goods" in ll or ll == "cogs":
            ps.cost_of_goods_sold = val
        elif ll == "gross margin":
            ps.gross_margin = val
        elif "promo" in ll and "expend" in ll:
            ps.promotional_expenditures = val
        elif "advertis" in ll and "expend" in ll:
            ps.advertising_expenditures = val
        elif ll == "product contribution":
            ps.product_contribution = val
        elif ll == "sales force":
            ps.sales_force_expense = val
        elif ll in ("admin", "administration"):
            ps.admin_expense = val
        elif ll == "fixed costs":
            ps.fixed_costs = val
        elif ll == "net income":
            ps.net_income = val
        elif ll == "cumulative net income":
            ps.cumulative_net_income = val
        elif ll == "stock price":
            ps.stock_price = val
        elif "marketing eff" in ll:
            ps.marketing_efficiency_index = val
        elif "capacity" in ll:
            ps.capacity_utilization_pct = val
        elif "unit" in ll and "share" in ll:
            ps.market_share_unit_pct = val
        elif "mfr" in ll and "share" in ll:
            ps.market_share_dollar_pct = val
        elif "retail" in ll and "share" in ll:
            # Also capture retail share
            if ps.market_share_dollar_pct is None:
                ps.market_share_dollar_pct = val
        elif "shelf space" in ll:
            ps.average_shelf_space_pct = val

    # Industry data in columns E-F
    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 5)
        if not label:
            continue
        ll = str(label).lower()
        val = _float(_cell(ws, r, 6))
        if "total mfr" in ll or "manufacturer sales" in ll:
            ps.industry_total_mfr_sales = val
        elif "unit sales" in ll:
            ps.industry_unit_sales = val
        elif "sales force" in ll:
            ps.industry_sales_force_spending = val
        elif "advertis" in ll:
            ps.industry_advertising_spending = val
        elif "digital" in ll:
            ps.industry_digital_marketing = val
        elif "promo" in ll:
            ps.industry_promotional_spending = val
        elif "net income" in ll:
            ps.industry_total_net_income = val

    wb.close()
    return ps


def parse_income_statement(year: int) -> IncomeStatement | None:
    wb = _open(year, "Income_Statement")
    if wb is None:
        return None
    ws = wb.active
    inc = IncomeStatement()

    row_map: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if label:
            row_map[str(label).strip()] = r

    def _amount(label):
        r = row_map.get(label)
        if r:
            # Try col C first (total), then B
            v = _float(_cell(ws, r, 3))
            if v is None:
                v = _float(_cell(ws, r, 2))
            return v
        return None

    def _ratio(label):
        r = row_map.get(label)
        if r:
            return _float(_cell(ws, r, 4))
        return None

    # Use fuzzy matching since labels have leading spaces and vary
    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if not label:
            continue
        ll = str(label).lower().strip()
        # Best amount: try col C, then B
        amt = _float(_cell(ws, r, 3))
        if amt is None:
            amt = _float(_cell(ws, r, 2))
        # Some sub-items are in col B only
        sub_amt = _float(_cell(ws, r, 2))
        ratio = _float(_cell(ws, r, 4))

        if "manufacturer" in ll and "sales" in ll:
            inc.manufacturer_sales = sub_amt or amt
            # ratio for mfr sales is always 1.0, skip
        elif "promotional allowance" in ll or "promo allowance" in ll:
            inc.promotional_allowance = sub_amt or amt
            inc.promotional_allowance_pct = ratio
        elif "cost of goods" in ll or ll == "cogs":
            inc.cost_of_goods_sold = sub_amt or amt
            inc.cogs_pct = ratio
        elif "gross margin" in ll:
            inc.gross_margin = amt
            inc.gross_margin_pct = ratio
        elif "consumer" in ll and "trade" in ll:
            inc.consumer_and_trade_promo = sub_amt or amt
            inc.consumer_trade_promo_pct = ratio
        elif ll.strip() == "advertising":
            inc.advertising = sub_amt or amt
            inc.advertising_pct = ratio
        elif "sales force" in ll:
            inc.sales_force = sub_amt or amt
            inc.sales_force_pct = ratio
        elif "admin" in ll:
            inc.admin = sub_amt or amt
            inc.admin_pct = ratio
        elif "total marketing" in ll:
            inc.total_marketing = amt
            inc.total_marketing_pct = ratio
        elif "contribution" in ll and "marketing" in ll:
            inc.contribution_after_marketing = amt
            inc.contribution_after_marketing_pct = ratio
        elif "fixed cost" in ll:
            inc.fixed_costs = amt
            inc.fixed_costs_pct = ratio
        elif ll.strip() == "net income":
            inc.net_income = amt
            inc.net_income_pct = ratio
        elif "budget" in ll:
            inc.next_year_budget = amt

    wb.close()
    return inc


def parse_product_contribution(year: int) -> list[ProductContribution]:
    wb = _open(year, "Product_Contribution")
    if wb is None:
        return []
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        pc = ProductContribution()
        # Sheet name is like "Allround - Overall" or "Allround - Per-Unit"
        parts = sheet_name.split(" - ")
        brand = parts[0].strip() if parts else sheet_name
        is_per_unit = "per" in sheet_name.lower()

        if is_per_unit:
            # Find matching brand in results
            existing = next((p for p in results if p.brand == brand), None)
            if existing is None:
                existing = ProductContribution(brand=brand)
                results.append(existing)
            target = existing
        else:
            target = ProductContribution(brand=brand)
            results.append(target)

        for r in range(1, ws.max_row + 1):
            label = _cell(ws, r, 1)
            if not label:
                continue
            ll = str(label).lower().strip()
            # Try col C (total) first, then B
            val = _float(_cell(ws, r, 3))
            if val is None:
                val = _float(_cell(ws, r, 2))

            if is_per_unit:
                if "avg retail" in ll:
                    target.per_unit_avg_retail_price = val
                elif "retail sales" in ll:
                    target.per_unit_retail_sales = val
                elif "mfr sales" in ll:
                    target.per_unit_manufacturer_sales = val
                elif "allowance" in ll:
                    target.per_unit_allowance_expense = val
                elif "cogs" in ll:
                    target.per_unit_cogs = val
                elif "gross margin" in ll:
                    target.per_unit_gross_margin = val
                elif "consumer" in ll and "trade" in ll:
                    target.per_unit_consumer_trade_promo = val
                elif "advertising" in ll:
                    target.per_unit_advertising = val
                elif "total marketing" in ll:
                    target.per_unit_total_marketing = val
                elif "product contribution" in ll:
                    target.per_unit_product_contribution = val
            else:
                if "unit sales" in ll:
                    target.unit_sales = val
                elif "avg retail" in ll:
                    target.avg_retail_price = val
                elif "retail sales" in ll:
                    target.retail_sales = val
                elif "mfr sales" in ll:
                    target.manufacturer_sales = val
                elif "allowance" in ll:
                    target.allowance_expense = val
                elif "cogs" in ll:
                    target.cost_of_goods_sold = val
                elif "gross margin" in ll:
                    target.gross_margin = val
                elif "consumer" in ll and "trade" in ll:
                    target.consumer_and_trade_promo = val
                elif "advertising" in ll:
                    target.advertising = val
                elif "total marketing" in ll:
                    target.total_marketing = val
                elif "product contribution" in ll:
                    target.product_contribution = val

    # Deduplicate: merge per-unit sheets into overall
    unique: dict[str, ProductContribution] = {}
    for pc in results:
        if pc.brand in unique:
            existing = unique[pc.brand]
            # Merge non-None per-unit fields
            for f in [
                "per_unit_avg_retail_price", "per_unit_retail_sales",
                "per_unit_manufacturer_sales", "per_unit_allowance_expense",
                "per_unit_cogs", "per_unit_gross_margin",
                "per_unit_consumer_trade_promo", "per_unit_advertising",
                "per_unit_total_marketing", "per_unit_product_contribution",
                "unit_sales", "avg_retail_price", "retail_sales",
                "manufacturer_sales", "allowance_expense", "cost_of_goods_sold",
                "gross_margin", "consumer_and_trade_promo", "advertising",
                "total_marketing", "product_contribution",
            ]:
                new_val = getattr(pc, f)
                if new_val is not None:
                    setattr(existing, f, new_val)
        else:
            unique[pc.brand] = pc
    wb.close()
    return list(unique.values())


def parse_sales_report(year: int) -> dict[str, BrandSalesReport]:
    wb = _open(year, "Sales_Report")
    if wb is None:
        return {}
    results = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        brand = sheet_name.strip()
        sr = BrandSalesReport(brand=brand)

        # Find sections by scanning for headers
        discount_start = None
        channel_start = None
        for r in range(1, ws.max_row + 1):
            label = _cell(ws, r, 1)
            if label and "discount" in str(label).lower():
                discount_start = r
            elif label and "distribution" in str(label).lower():
                channel_start = r

        # Parse discount schedule section
        if discount_start:
            tier_map = {
                "<250": "direct_under_250",
                "<2500": "direct_under_2500",
                "2500+": "direct_2500_plus",
            }
            for r in range(discount_start + 1, (channel_start or ws.max_row) + 1):
                label = _cell(ws, r, 1)
                if not label:
                    continue
                ls = str(label).strip()
                for key, attr in tier_map.items():
                    if key in ls:
                        setattr(sr, attr, DiscountTierSales(
                            price=_float(_cell(ws, r, 2)),
                            units=_float(_cell(ws, r, 3)),
                            dollars=_float(_cell(ws, r, 4)),
                            pct_of_total=_float(_cell(ws, r, 5)),
                        ))
                if "wholesale" in ls.lower() or "indirect" in ls.lower():
                    sr.indirect_wholesale = DiscountTierSales(
                        price=_float(_cell(ws, r, 2)),
                        units=_float(_cell(ws, r, 3)),
                        dollars=_float(_cell(ws, r, 4)),
                        pct_of_total=_float(_cell(ws, r, 5)),
                    )

        # Parse channel section
        if channel_start:
            channel_map = {
                "indep": "channel_independent_drugstores",
                "chain": "channel_chain_drugstores",
                "grocery": "channel_grocery_stores",
                "convenience": "channel_convenience_stores",
                "mass": "channel_mass_merchandisers",
                "wholesale": "channel_wholesalers",
            }
            for r in range(channel_start + 1, ws.max_row + 1):
                label = _cell(ws, r, 1)
                if not label:
                    continue
                ls = str(label).lower()
                for key, attr in channel_map.items():
                    if key in ls:
                        setattr(sr, attr, ChannelSalesDetail(
                            sales_force_count=_float(_cell(ws, r, 2)),
                            units=_float(_cell(ws, r, 3)),
                            dollars=_float(_cell(ws, r, 4)),
                            pct_of_total=_float(_cell(ws, r, 5)),
                        ))
                        break

        results[brand] = sr
    wb.close()
    return results


def parse_promotion_report(year: int) -> dict[str, BrandPromotionReport]:
    wb = _open(year, "Promotion_Report")
    if wb is None:
        return {}
    results = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        brand = sheet_name.strip()
        pr = BrandPromotionReport(brand=brand)

        for r in range(1, ws.max_row + 1):
            label = _cell(ws, r, 1)
            if not label:
                continue
            ll = str(label).lower()
            val = _float(_cell(ws, r, 2))
            if val is None:
                val = _float(_cell(ws, r, 3))

            if "promo allowance" in ll:
                pr.promotional_allowance = val
            elif "co-op" in ll:
                pr.coop_advertising = val
            elif "pop" in ll or "point of purchase" in ll:
                pr.point_of_purchase = val
            elif "trial" in ll:
                pr.trial_size = val
            elif "expir" in ll:
                pr.coupon_expiration_months = val
            elif "coupon amount" in ll:
                pr.coupon_amount = val
            elif "trade rating" in ll:
                pr.trade_rating = val
            elif "participating" in ll:
                pr.pct_participating_retailers = val
            elif "mailed" in ll:
                pr.coupons_mailed_thousands = val
            elif "conversion" in ll:
                pr.coupon_conversions = val
            elif "redeemed" in ll:
                pr.coupons_redeemed = val
            elif "% of total" in ll:
                pr.promo_pct_of_total_sales = val

        results[brand] = pr
    wb.close()
    return results


def parse_dashboard(year: int) -> Dashboard | None:
    wb = _open(year, "Dashboard")
    if wb is None:
        return None
    ws = wb.active
    d = Dashboard()

    # Find the data table rows
    metrics = [
        ("stock price", "stock_price"),
        ("unit sales", "unit_sales"),
        ("revenue", "revenue"),
        ("net income", "net_income"),
        ("cumulative", "cumulative_net_income"),
    ]

    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if not label:
            continue
        ll = str(label).lower()
        for key, attr in metrics:
            if key in ll and "cumulative" not in ll if key != "cumulative" else key in ll:
                setattr(d, f"{attr}_previous", _float(_cell(ws, r, 2)))
                setattr(d, f"{attr}_current", _float(_cell(ws, r, 3)))
                setattr(d, f"{attr}_change", _float(_cell(ws, r, 4)))
                break

        if "market update" in ll:
            texts = []
            for mr in range(r + 1, min(r + 10, ws.max_row + 1)):
                t = _cell(ws, mr, 1)
                if t and "company" not in str(t).lower():
                    texts.append(str(t))
                else:
                    break
            d.market_update_text = " ".join(texts)

        if "company message" in ll:
            texts = []
            for mr in range(r + 1, min(r + 10, ws.max_row + 1)):
                t = _cell(ws, mr, 1)
                if t:
                    texts.append(str(t))
            d.company_messages_text = " ".join(texts)

    wb.close()
    return d


def parse_brand_formulations(year: int) -> dict[str, BrandFormulation]:
    wb = _open(year, "Brand_Formulations")
    if wb is None:
        return {}
    ws = wb.active
    results = {}

    # Find header row and data rows
    header_row = None
    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if label and "brand" in str(label).lower():
            header_row = r
            break

    if header_row is None:
        wb.close()
        return {}

    # Skip the max-allowed row (usually header_row + 1)
    data_start = header_row + 1
    # Check if next row has "Max" or numeric max values
    max_label = _cell(ws, data_start, 1)
    if max_label and ("max" in str(max_label).lower() or _float(max_label) is not None):
        data_start += 1

    skip_keywords = ["brand formulation", "competitor", "all comp", "note"]
    for r in range(data_start, ws.max_row + 1):
        brand = _cell(ws, r, 1)
        if not brand or str(brand).strip() == "":
            continue
        brand = str(brand).strip()
        if any(kw in brand.lower() for kw in skip_keywords):
            continue
        # Must have at least one numeric ingredient value
        if _float(_cell(ws, r, 2)) is None and _float(_cell(ws, r, 3)) is None:
            continue
        bf = BrandFormulation(
            brand=brand,
            analgesic_mg=_float(_cell(ws, r, 2)),
            antihistamine_mg=_float(_cell(ws, r, 3)),
            decongestant_mg=_float(_cell(ws, r, 4)),
            cough_suppressant_mg=_float(_cell(ws, r, 5)),
            expectorant_mg=_float(_cell(ws, r, 6)),
            alcohol_pct=_float(_cell(ws, r, 7)),
            description=str(_cell(ws, r, 8) or ""),
        )
        results[brand] = bf

    wb.close()
    return results


def parse_symptoms_reported(year: int) -> SymptomsReported | None:
    wb = _open(year, "Symptoms_Reported")
    if wb is None:
        return None
    ws = wb.active
    sr = SymptomsReported()

    symptom_map = {
        "aches": "aches_pct",
        "nasal": "nasal_congestion_pct",
        "chest": "chest_congestion_pct",
        "runny": "runny_nose_pct",
        "cough": "coughing_pct",
        "allergy": "allergy_symptoms_pct",
    }

    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if not label:
            continue
        ll = str(label).lower()
        for key, attr in symptom_map.items():
            if key in ll:
                setattr(sr, attr, _float(_cell(ws, r, 2)))
                break

    wb.close()
    return sr


def parse_industry_outlook(year: int) -> IndustryOutlook | None:
    wb = _open(year, "Industry_Outlook")
    if wb is None:
        return None
    ws = wb.active
    io = IndustryOutlook()

    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if not label:
            continue
        ll = str(label).lower()

        if "pop" in ll and "growth" in ll:
            io.population_growth_pct = _float(_cell(ws, r, 2))
            io.population_growth_forecast = str(_cell(ws, r, 3) or "")
        elif "industry" in ll and "growth" in ll:
            io.industry_growth_pct = _float(_cell(ws, r, 2))
            io.industry_growth_forecast = str(_cell(ws, r, 3) or "")
        elif "inflation" in ll:
            io.inflation_rate_pct = _float(_cell(ws, r, 2))
            io.inflation_rate_forecast = str(_cell(ws, r, 3) or "")
        elif "direct" in ll and "%" in ll:
            io.pct_direct_sales = _float(_cell(ws, r, 2))
        elif "wholesale" in ll and "%" in ll:
            io.pct_wholesale_sales = _float(_cell(ws, r, 2))
        elif "salary" in ll:
            io.sales_rep_salary = _float(_cell(ws, r, 2))
        elif "expenses" in ll and "sales" not in ll:
            io.sales_rep_expenses = _float(_cell(ws, r, 2))
        elif "training" in ll:
            io.sales_rep_training = _float(_cell(ws, r, 2))
        else:
            # Marketing research cost items
            val = _float(_cell(ws, r, 2))
            if val is not None and val > 0 and "$" not in ll:
                io.research_costs[str(label).strip()] = val

    wb.close()
    return io


def parse_advertising(year: int) -> dict[str, BrandAdvertising]:
    wb = _open(year, "Advertising")
    if wb is None:
        return {}
    ws = wb.active
    results = {}

    # Find header row - look for "Media" or "Expend" or "Primary" in any cell
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(2, ws.max_column + 1):
            v = _cell(ws, r, c)
            if v and ("media" in str(v).lower() or "primary" in str(v).lower()):
                header_row = r
                break
        if header_row:
            break

    if header_row is None:
        wb.close()
        return {}

    for r in range(header_row + 1, ws.max_row + 1):
        brand = _cell(ws, r, 1)
        if not brand or str(brand).strip() == "" or "note" in str(brand).lower():
            continue
        brand = str(brand).strip()
        ba = BrandAdvertising(
            brand=brand,
            media_expenditure=_float(_cell(ws, r, 2)),
            ad_agency=str(_cell(ws, r, 3) or ""),
            primary_pct=_float(_cell(ws, r, 4)),
            benefits_pct=_float(_cell(ws, r, 5)),
            comparison_pct=_float(_cell(ws, r, 6)),
            reminder_pct=_float(_cell(ws, r, 7)),
        )
        results[brand] = ba

    wb.close()
    return results


def _parse_brand_survey(year: int, filename: str, value_cols: list[str]) -> dict[str, dict[str, Any]]:
    """Generic parser for brand survey sheets (perceptions, awareness, etc.)."""
    wb = _open(year, filename)
    if wb is None:
        return {}
    ws = wb.active
    results: dict[str, dict[str, Any]] = {}

    # Find header row - look for column headers
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(2, ws.max_column + 1):
            v = _cell(ws, r, c)
            if v and str(v).strip() in value_cols:
                header_row = r
                break
        if header_row:
            break

    if header_row is None:
        # Try finding any row with multiple values
        for r in range(1, min(10, ws.max_row + 1)):
            vals = [_cell(ws, r, c) for c in range(2, ws.max_column + 1)]
            non_none = [v for v in vals if v is not None]
            if len(non_none) >= 2:
                header_row = r
                break

    if header_row is None:
        wb.close()
        return {}

    # Map column indices to header names
    col_map: dict[int, str] = {}
    for c in range(2, ws.max_column + 1):
        v = _cell(ws, header_row, c)
        if v:
            col_map[c] = str(v).strip()

    # Parse brand rows - skip rows without any numeric data in value columns
    for r in range(header_row + 1, ws.max_row + 1):
        brand = _cell(ws, r, 1)
        if not brand or str(brand).strip() == "":
            continue
        brand = str(brand).strip()
        # Skip note/scale rows (e.g. "1 = Not at all effective, 5 = ...")
        if "=" in brand and any(c.isdigit() for c in brand[:3]):
            continue
        if brand.lower().startswith("note"):
            continue
        row_data: dict[str, Any] = {}
        has_numeric = False
        for c, header in col_map.items():
            v = _cell(ws, r, c)
            row_data[header] = v
            if _float(v) is not None:
                has_numeric = True
        if not has_numeric:
            continue
        results[brand] = row_data

    wb.close()
    return results


def parse_brand_perceptions(year: int) -> dict[str, BrandPerception]:
    raw = _parse_brand_survey(year, "Brand_Perceptions",
        ["Aches", "Nasal Cong.", "Chest Cong.", "Runny Nose", "Coughing", "Allergies"])
    results = {}
    for brand, data in raw.items():
        bp = BrandPerception(brand=brand)
        for key, val in data.items():
            kl = key.lower()
            if "ache" in kl:
                bp.aches = _float(val)
            elif "nasal" in kl:
                bp.nasal_congestion = _float(val)
            elif "chest" in kl:
                bp.chest_congestion = _float(val)
            elif "runny" in kl:
                bp.runny_nose = _float(val)
            elif "cough" in kl:
                bp.coughing = _float(val)
            elif "allerg" in kl:
                bp.allergies = _float(val)
        results[brand] = bp
    return results


def parse_brand_awareness(year: int) -> dict[str, BrandAwareness]:
    raw = _parse_brand_survey(year, "Brand_Awareness",
        ["Brand Awareness", "Brand Trials", "Most Freq. Purchase", "Conversion Ratio", "Retention Ratio"])
    results = {}
    for brand, data in raw.items():
        ba = BrandAwareness(brand=brand)
        for key, val in data.items():
            kl = key.lower()
            if "awareness" in kl:
                ba.brand_awareness_pct = _float(val)
            elif "trial" in kl:
                ba.brand_trial_pct = _float(val)
            elif "freq" in kl:
                ba.most_frequent_purchase_pct = _float(val)
            elif "conversion" in kl:
                ba.conversion_ratio = _float(val)
            elif "retention" in kl:
                ba.retention_ratio = _float(val)
        results[brand] = ba
    return results


def parse_shelf_space(year: int) -> dict[str, BrandShelfSpace]:
    wb = _open(year, "Shelf_Space")
    if wb is None:
        return {}
    ws = wb.active
    results = {}

    # Find header row
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(2, ws.max_column + 1):
            v = _cell(ws, r, c)
            if v and "drug" in str(v).lower():
                header_row = r
                break
        if header_row:
            break

    if header_row is None:
        wb.close()
        return {}

    for r in range(header_row + 1, ws.max_row + 1):
        brand = _cell(ws, r, 1)
        if not brand or str(brand).strip() == "":
            continue
        brand = str(brand).strip()
        bs = BrandShelfSpace(
            brand=brand,
            independent_drugstores=_float(_cell(ws, r, 2)),
            chain_drugstores=_float(_cell(ws, r, 3)),
            grocery_stores=_float(_cell(ws, r, 4)),
            convenience_stores=_float(_cell(ws, r, 5)),
            mass_merchandisers=_float(_cell(ws, r, 6)),
            all_channels=_float(_cell(ws, r, 7)),
        )
        results[brand] = bs

    wb.close()
    return results


def parse_promotion(year: int) -> dict[str, BrandPromotion]:
    wb = _open(year, "Promotion")
    if wb is None:
        return {}
    ws = wb.active
    results = {}

    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(2, ws.max_column + 1):
            v = _cell(ws, r, c)
            if v and "allow" in str(v).lower():
                header_row = r
                break
        if header_row:
            break

    if header_row is None:
        wb.close()
        return {}

    for r in range(header_row + 1, ws.max_row + 1):
        brand = _cell(ws, r, 1)
        if not brand or str(brand).strip() == "":
            continue
        brand = str(brand).strip()
        if "note" in brand.lower() or "dollar" in brand.lower():
            continue
        bp = BrandPromotion(
            brand=brand,
            promotional_allowance_pct=_float(_cell(ws, r, 2)),
            coop_advertising=_float(_cell(ws, r, 3)),
            point_of_purchase=_float(_cell(ws, r, 4)),
            trial_size=str(_cell(ws, r, 5) or ""),
            coupon_amount=_float(_cell(ws, r, 6)),
            trade_rating=_float(_cell(ws, r, 7)),
        )
        results[brand] = bp

    wb.close()
    return results


def parse_conjoint_analysis(year: int) -> ConjointAnalysis | None:
    wb = _open(year, "Conjoint_Analysis")
    if wb is None:
        return None
    ws = wb.active
    ca = ConjointAnalysis()

    # Find "Rank Products @" row - price points are in that row
    rank_row = None
    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if label and "rank" in str(label).lower():
            rank_row = r
            break

    if rank_row:
        # Price points are in cols C onwards (col B may be empty)
        for c in range(2, ws.max_column + 1):
            v = _float(_cell(ws, rank_row, c))
            if v is not None:
                ca.price_points.append(v)

        # Product ranking rows follow immediately
        for r in range(rank_row + 1, ws.max_row + 1):
            label = _cell(ws, r, 1)
            if not label:
                break
            ls = str(label).strip()
            rankings = []
            for c in range(2, ws.max_column + 1):
                v = _float(_cell(ws, r, c))
                if v is not None:
                    rankings.append(int(v))
            if rankings:
                ca.product_rankings[ls] = rankings

    # Parse utility section - find the "Product Utility" / "Price Utility" header row
    import re
    utility_header_row = None
    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if label and "product utility" in str(label).lower():
            utility_header_row = r
            m = re.search(r'([\d.]+)%', str(label))
            if m:
                ca.product_importance_pct = float(m.group(1))
            # Price utility % may be in col E of same row
            c5 = _cell(ws, r, 5)
            if c5 and "price utility" in str(c5).lower():
                m2 = re.search(r'([\d.]+)%', str(c5))
                if m2:
                    ca.price_importance_pct = float(m2.group(1))
            break

    if utility_header_row:
        # Product utilities in rows after header: col A = name, col C = utility
        # Price utilities in same rows: col E = price, col G = utility
        ca.price_utilities = {}  # Reset to avoid stale data
        for r in range(utility_header_row + 1, ws.max_row + 1):
            label = _cell(ws, r, 1)
            if label:
                ls = str(label).strip()
                util_val = _float(_cell(ws, r, 3))
                if util_val is not None:
                    ca.product_utilities[ls] = util_val

            # Price utility in cols E and G
            price = _float(_cell(ws, r, 5))
            util = _float(_cell(ws, r, 7))
            if price is not None and util is not None:
                ca.price_utilities[price] = util

    wb.close()
    return ca


def parse_satisfaction(year: int) -> dict[str, BrandSatisfaction]:
    raw = _parse_brand_survey(year, "Satisfaction",
        ["Cross Section", "Overall"])
    results = {}
    for brand, data in raw.items():
        bs = BrandSatisfaction(brand=brand)
        for key, val in data.items():
            kl = key.lower()
            if "cross" in kl:
                bs.cross_section = _float(val)
            elif "overall" in kl:
                bs.overall = _float(val)
        results[brand] = bs
    return results


def parse_sales_force(year: int) -> dict[str, CompanySalesForce]:
    wb = _open(year, "Sales_Force")
    if wb is None:
        return {}
    results: dict[str, CompanySalesForce] = {}

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        is_pct = "percent" in sheet_name.lower()

        # Find header row with company names
        header_row = None
        col_map: dict[int, str] = {}
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(2, ws.max_column + 1):
                v = _cell(ws, r, c)
                if v and str(v).strip() in COMPANIES:
                    header_row = r
                    break
            if header_row:
                break

        if header_row is None:
            continue

        for c in range(2, ws.max_column + 1):
            v = _cell(ws, header_row, c)
            if v and str(v).strip() in COMPANIES:
                col_map[c] = str(v).strip()
                if str(v).strip() not in results:
                    results[str(v).strip()] = CompanySalesForce(company=str(v).strip())

        # Direct SF row labels
        direct_fields = [
            ("indep", "independent_drugstores"),
            ("chain", "chain_drugstores"),
            ("grocery", "grocery_stores"),
            ("convenience", "convenience_stores"),
            ("mass", "mass_merchandisers"),
        ]
        indirect_fields = [
            ("wholesaler", "wholesaler_support"),
            ("merchandis", "merchandisers"),
            ("detailer", "detailers"),
        ]

        for r in range(header_row + 1, ws.max_row + 1):
            label = _cell(ws, r, 1)
            if not label:
                continue
            ll = str(label).lower()

            for c, company in col_map.items():
                csf = results[company]
                val = _float(_cell(ws, r, c))

                if is_pct:
                    for key, fname in direct_fields:
                        if key in ll:
                            setattr(csf, f"pct_{fname}", val)
                            break
                    for key, fname in indirect_fields:
                        if key in ll:
                            setattr(csf, f"pct_{fname}", val)
                            break
                    if "total direct" in ll:
                        csf.pct_total_direct = val
                    elif "total indirect" in ll:
                        csf.pct_total_indirect = val
                else:
                    for key, fname in direct_fields:
                        if key in ll:
                            setattr(csf, f"direct_{fname}", val)
                            break
                    for key, fname in indirect_fields:
                        if key in ll:
                            setattr(csf, f"indirect_{fname}", val)
                            break
                    if "total direct" in ll:
                        csf.total_direct = val
                    elif "total indirect" in ll:
                        csf.total_indirect = val
                    elif "total" in ll and "sf" in ll:
                        csf.total_sales_force = val

    wb.close()
    return results


def parse_pricing(year: int) -> tuple[dict[str, BrandPricing], dict[str, BrandChannelDiscountDetail]]:
    wb = _open(year, "Pricing")
    if wb is None:
        return {}, {}
    pricing_results: dict[str, BrandPricing] = {}
    discount_results: dict[str, BrandChannelDiscountDetail] = {}

    # First sheet: Pricing
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if "discount" in sheet_name.lower():
            # Channel discount detail sheet
            current_brand = None
            current_detail: BrandChannelDiscountDetail | None = None
            for r in range(1, ws.max_row + 1):
                a_val = _cell(ws, r, 1)
                if a_val and str(a_val).strip() and not any(
                    kw in str(a_val).lower() for kw in ["discount", "total", "volume", "indep", "chain", "grocery", "convenience", "mass", "note", "%"]
                ):
                    # Brand name row
                    if current_detail and current_brand:
                        discount_results[current_brand] = current_detail
                    current_brand = str(a_val).strip()
                    current_detail = BrandChannelDiscountDetail(brand=current_brand)
                elif a_val and "%" in str(a_val) and current_detail is not None:
                    # Discount tier row
                    tier = ChannelDiscountTier(
                        discount_pct=_float(str(a_val).replace("%", "")),
                        units_independent_drugstores=_float(_cell(ws, r, 2)),
                        units_chain_drugstores=_float(_cell(ws, r, 3)),
                        units_grocery_stores=_float(_cell(ws, r, 4)),
                        units_convenience_stores=_float(_cell(ws, r, 5)),
                        units_mass_merchandisers=_float(_cell(ws, r, 6)),
                    )
                    current_detail.tiers.append(tier)
            if current_detail and current_brand:
                discount_results[current_brand] = current_detail
        else:
            # Main pricing sheet
            header_row = None
            for r in range(1, min(10, ws.max_row + 1)):
                for c in range(2, ws.max_column + 1):
                    v = _cell(ws, r, c)
                    if v and ("msrp" in str(v).lower() or "indep" in str(v).lower()):
                        header_row = r
                        break
                if header_row:
                    break

            if header_row is None:
                continue

            for r in range(header_row + 1, ws.max_row + 1):
                brand = _cell(ws, r, 1)
                if not brand or str(brand).strip() == "":
                    continue
                brand = str(brand).strip()
                bp = BrandPricing(
                    brand=brand,
                    msrp=_float(_cell(ws, r, 2)),
                    avg_retail_independent_drugstores=_float(_cell(ws, r, 3)),
                    avg_retail_chain_drugstores=_float(_cell(ws, r, 4)),
                    avg_retail_grocery_stores=_float(_cell(ws, r, 5)),
                    avg_retail_convenience_stores=_float(_cell(ws, r, 6)),
                    avg_retail_mass_merchandisers=_float(_cell(ws, r, 7)),
                )
                pricing_results[brand] = bp

    wb.close()
    return pricing_results, discount_results


def parse_decision_criteria(year: int) -> DecisionCriteria | None:
    wb = _open(year, "Decision_Criteria")
    if wb is None:
        return None
    ws = wb.active
    dc = DecisionCriteria()

    criteria_map = {
        "effectiveness": "effectiveness_rankings",
        "side effect": "side_effects_rankings",
        "price": "price_rankings",
        "form": "form_rankings",
        "duration": "duration_rankings",
    }

    for r in range(1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if not label:
            continue
        ll = str(label).lower()

        if "penetration" in ll:
            dc.market_penetration_pct = _float(_cell(ws, r, 2))
        elif "purchase" in ll and "year" in ll:
            dc.avg_purchase_per_year = _float(_cell(ws, r, 2))
        else:
            for key, attr in criteria_map.items():
                if key in ll:
                    rankings = []
                    for c in range(2, 7):
                        v = _float(_cell(ws, r, c))
                        if v is not None:
                            rankings.append(v)
                    setattr(dc, attr, rankings)
                    break

    wb.close()
    return dc


def parse_recommendations(year: int) -> dict[str, BrandRecommendation]:
    wb = _open(year, "Recommendations")
    if wb is None:
        return {}
    ws = wb.active
    results = {}

    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(2, ws.max_column + 1):
            v = _cell(ws, r, c)
            if v and "cold" in str(v).lower():
                header_row = r
                break
        if header_row:
            break

    if header_row is None:
        wb.close()
        return {}

    for r in range(header_row + 1, ws.max_row + 1):
        brand = _cell(ws, r, 1)
        if not brand or str(brand).strip() == "":
            continue
        brand = str(brand).strip()
        br = BrandRecommendation(
            brand=brand,
            cold_pct=_float(_cell(ws, r, 2)),
            cough_pct=_float(_cell(ws, r, 3)),
            allergy_pct=_float(_cell(ws, r, 4)),
        )
        results[brand] = br

    wb.close()
    return results


def parse_operating_statistics(year: int) -> dict[str, CompanyOperatingStats]:
    wb = _open(year, "Operating_Statistics")
    if wb is None:
        return {}
    results: dict[str, CompanyOperatingStats] = {}

    row_fields_abs = [
        ("retail sales", "retail_sales"),
        ("manufacturer sales", "manufacturer_sales"),
        ("promotional allowance", "promotional_allowance"),
        ("cost of goods", "cost_of_goods_sold"),
        ("gross margin", "gross_margin"),
        ("cons.", "consumer_and_trade_promo"),
        ("consumer", "consumer_and_trade_promo"),
        ("advertising", "advertising"),
        ("sales force", "sales_force"),
        ("admin", "admin"),
        ("contrib", "contribution_after_marketing"),
        ("fixed cost", "fixed_costs"),
        ("net income", "net_income"),
        ("stock price", "stock_price"),
        ("capacity", "capacity_utilization_pct"),
    ]

    row_fields_pct_retail = [
        ("mfr sales", "pct_retail_manufacturer_sales"),
        ("promo allowance", "pct_retail_promo_allowance"),
        ("cogs", "pct_retail_cogs"),
        ("gross margin", "pct_retail_gross_margin"),
        ("consumer", "pct_retail_consumer_trade_promo"),
        ("advertising", "pct_retail_advertising"),
        ("sales force", "pct_retail_sales_force"),
        ("admin", "pct_retail_admin"),
        ("contrib", "pct_retail_contrib_after_marketing"),
        ("fixed cost", "pct_retail_fixed_costs"),
        ("net income", "pct_retail_net_income"),
    ]

    row_fields_pct_mfr = [
        ("promo allowance", "pct_mfr_promo_allowance"),
        ("cogs", "pct_mfr_cogs"),
        ("gross margin", "pct_mfr_gross_margin"),
        ("consumer", "pct_mfr_consumer_trade_promo"),
        ("advertising", "pct_mfr_advertising"),
        ("sales force", "pct_mfr_sales_force"),
        ("admin", "pct_mfr_admin"),
        ("contrib", "pct_mfr_contrib_after_marketing"),
        ("fixed cost", "pct_mfr_fixed_costs"),
        ("net income", "pct_mfr_net_income"),
    ]

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        sn_lower = sheet_name.lower()

        if sheet_idx == 0:
            field_map = row_fields_abs
        elif "retail" in sn_lower:
            field_map = row_fields_pct_retail
        elif "manufacturer" in sn_lower or "mfr" in sn_lower:
            field_map = row_fields_pct_mfr
        else:
            # Guess by sheet index
            field_map = [row_fields_abs, row_fields_pct_retail, row_fields_pct_mfr][min(sheet_idx, 2)]

        # Find header row with company names
        header_row = None
        col_map: dict[int, str] = {}
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(2, ws.max_column + 1):
                v = _cell(ws, r, c)
                if v and str(v).strip() in COMPANIES:
                    header_row = r
                    break
            if header_row:
                break

        if header_row is None:
            continue

        for c in range(2, ws.max_column + 1):
            v = _cell(ws, header_row, c)
            if v and str(v).strip() in COMPANIES:
                col_map[c] = str(v).strip()
                if str(v).strip() not in results:
                    results[str(v).strip()] = CompanyOperatingStats(company=str(v).strip())

        for r in range(header_row + 1, ws.max_row + 1):
            label = _cell(ws, r, 1)
            if not label:
                continue
            ll = str(label).lower()
            if "note" in ll:
                continue

            for c, company in col_map.items():
                cos = results[company]
                val = _float(_cell(ws, r, c))

                for key, attr in field_map:
                    if key in ll:
                        setattr(cos, attr, val)
                        break

    wb.close()
    return results


def parse_manufacturer_sales(year: int) -> dict[str, BrandManufacturerSales]:
    wb = _open(year, "Manufacturer_Sales")
    if wb is None:
        return {}
    results: dict[str, BrandManufacturerSales] = {}

    segments = ["cold", "cough", "allergy", "nasal"]

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        is_share = "share" in sheet_name.lower() or sheet_idx == 0

        # Find header row with segment names
        header_row = None
        col_map: dict[int, str] = {}
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(2, ws.max_column + 1):
                v = _cell(ws, r, c)
                if v and any(s in str(v).lower() for s in segments):
                    header_row = r
                    break
            if header_row:
                break

        if header_row is None:
            continue

        for c in range(2, ws.max_column + 1):
            v = _cell(ws, header_row, c)
            if v:
                vs = str(v).lower()
                for s in segments + ["total"]:
                    if s in vs:
                        col_map[c] = s
                        break

        skip_kw = ["total", "mfr", "growth", "manufacturer", "market share", "note", "sales"]
        for r in range(header_row + 1, ws.max_row + 1):
            brand = _cell(ws, r, 1)
            if not brand or str(brand).strip() == "":
                continue
            brand = str(brand).strip()
            if any(kw in brand.lower() for kw in skip_kw):
                continue
            if brand not in results:
                results[brand] = BrandManufacturerSales(brand=brand)
            bms = results[brand]

            for c, seg in col_map.items():
                val = _float(_cell(ws, r, c))
                if is_share:
                    if seg == "cold":
                        bms.cold_share_pct = val
                    elif seg == "cough":
                        bms.cough_share_pct = val
                    elif seg == "allergy":
                        bms.allergy_share_pct = val
                    elif seg == "nasal":
                        bms.nasal_spray_share_pct = val
                    elif seg == "total":
                        bms.total_share_pct = val
                else:
                    if seg == "cold":
                        bms.cold_sales = val
                    elif seg == "cough":
                        bms.cough_sales = val
                    elif seg == "allergy":
                        bms.allergy_sales = val
                    elif seg == "nasal":
                        bms.nasal_spray_sales = val
                    elif seg == "total":
                        bms.total_sales = val

    wb.close()
    return results


def parse_channel_sales(year: int) -> dict[str, BrandChannelSales]:
    wb = _open(year, "Channel_Sales")
    if wb is None:
        return {}
    results: dict[str, BrandChannelSales] = {}

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        sn_lower = sheet_name.lower()
        if "discount" in sn_lower:
            continue  # Skip channel discount detail (handled in pricing)

        # Determine if this is dollar amounts or percentages
        is_pct = sheet_idx == 1 or "%" in sn_lower or "percent" in sn_lower

        # Find header row
        header_row = None
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(2, ws.max_column + 1):
                v = _cell(ws, r, c)
                if v and ("drug" in str(v).lower() or "indep" in str(v).lower()):
                    header_row = r
                    break
            if header_row:
                break

        if header_row is None:
            continue

        skip_kw = ["total", "retail", "growth", "market share", "note", "sales"]
        for r in range(header_row + 1, ws.max_row + 1):
            brand = _cell(ws, r, 1)
            if not brand or str(brand).strip() == "":
                continue
            brand = str(brand).strip()
            if any(kw in brand.lower() for kw in skip_kw):
                continue
            if brand not in results:
                results[brand] = BrandChannelSales(brand=brand)
            bcs = results[brand]

            if is_pct:
                bcs.independent_drugstores_share_pct = _float(_cell(ws, r, 2))
                bcs.chain_drugstores_share_pct = _float(_cell(ws, r, 3))
                bcs.grocery_stores_share_pct = _float(_cell(ws, r, 4))
                bcs.convenience_stores_share_pct = _float(_cell(ws, r, 5))
                bcs.mass_merchandisers_share_pct = _float(_cell(ws, r, 6))
                bcs.total_share_pct = _float(_cell(ws, r, 7))
            else:
                bcs.independent_drugstores_sales = _float(_cell(ws, r, 2))
                bcs.chain_drugstores_sales = _float(_cell(ws, r, 3))
                bcs.grocery_stores_sales = _float(_cell(ws, r, 4))
                bcs.convenience_stores_sales = _float(_cell(ws, r, 5))
                bcs.mass_merchandisers_sales = _float(_cell(ws, r, 6))
                bcs.total_sales = _float(_cell(ws, r, 7))

    wb.close()
    return results


def parse_purchase_intentions(year: int) -> dict[str, BrandPurchaseIntention]:
    raw = _parse_brand_survey(year, "Purchase_Intentions",
        ["Intended", "Bought"])
    results = {}
    for brand, data in raw.items():
        bpi = BrandPurchaseIntention(brand=brand)
        for key, val in data.items():
            kl = key.lower()
            if "intend" in kl:
                bpi.intended_pct = _float(val)
            elif "bought" in kl:
                bpi.bought_pct = _float(val)
        results[brand] = bpi
    return results


def parse_shopping_habits(year: int) -> ShoppingHabits | None:
    wb = _open(year, "Shopping_Habits")
    if wb is None:
        return None
    ws = wb.active
    sh = ShoppingHabits()

    # Find header row
    header_row = None
    col_map: dict[int, str] = {}
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(2, ws.max_column + 1):
            v = _cell(ws, r, c)
            if v and "cold" in str(v).lower():
                header_row = r
                break
        if header_row:
            break

    if header_row is None:
        wb.close()
        return sh

    for c in range(2, ws.max_column + 1):
        v = _cell(ws, header_row, c)
        if v:
            col_map[c] = str(v).strip().lower()

    channel_names = [
        ("indep", "Independent Drugstores"),
        ("chain", "Chain Drugstores"),
        ("grocery", "Grocery Stores"),
        ("convenience", "Convenience Stores"),
        ("mass", "Mass Merchandisers"),
    ]

    for r in range(header_row + 1, ws.max_row + 1):
        label = _cell(ws, r, 1)
        if not label:
            continue
        ls = str(label).strip()
        ll = ls.lower()

        channel = None
        for key, name in channel_names:
            if key in ll:
                channel = name
                break

        if channel is None:
            continue

        for c, seg in col_map.items():
            val = _float(_cell(ws, r, c))
            if "cold" in seg:
                sh.cold[channel] = val or 0.0
            elif "cough" in seg:
                sh.cough[channel] = val or 0.0
            elif "allergy" in seg:
                sh.allergy[channel] = val or 0.0

    wb.close()
    return sh


def parse_trade_offs(year: int) -> dict[str, BrandTradeOff]:
    raw = _parse_brand_survey(year, "Trade_Offs",
        ["MSRP", "Perceived Price", "Perceived Effect", "Purchased"])
    results = {}
    for brand, data in raw.items():
        bt = BrandTradeOff(brand=brand)
        for key, val in data.items():
            kl = key.lower()
            if "msrp" in kl:
                bt.msrp = _float(val)
            elif "price" in kl and "perceived" in kl:
                bt.perceived_price = str(val or "")
            elif "effect" in kl:
                bt.perceived_effectiveness = str(val or "")
            elif "purchased" in kl:
                bt.purchased_pct = _float(val)
        results[brand] = bt
    return results


def parse_brands_purchased(year: int) -> tuple[dict[str, BrandPurchaseSummary], dict[str, BrandPurchaseDetail]]:
    wb = _open(year, "Brands_Purchased")
    if wb is None:
        return {}, {}

    summaries: dict[str, BrandPurchaseSummary] = {}
    details: dict[str, BrandPurchaseDetail] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sn_lower = sheet_name.lower()

        if "survey" in sn_lower or "marketing" in sn_lower:
            # Summary sheet - same structure as other surveys
            header_row = None
            for r in range(1, min(10, ws.max_row + 1)):
                for c in range(2, ws.max_column + 1):
                    v = _cell(ws, r, c)
                    if v and ("cross" in str(v).lower() or "overall" in str(v).lower()):
                        header_row = r
                        break
                if header_row:
                    break

            if header_row is None:
                continue

            for r in range(header_row + 1, ws.max_row + 1):
                brand = _cell(ws, r, 1)
                if not brand or str(brand).strip() == "":
                    continue
                brand = str(brand).strip()
                bps = BrandPurchaseSummary(brand=brand)
                bps.cross_section_pct = _float(_cell(ws, r, 2))
                bps.overall_pct = _float(_cell(ws, r, 3))
                summaries[brand] = bps

        else:
            # Detail sheet - repeating blocks per brand
            current_brand = None
            current_detail: BrandPurchaseDetail | None = None

            for r in range(1, ws.max_row + 1):
                a_val = _cell(ws, r, 1)
                b_val = _cell(ws, r, 2)
                c_val = _cell(ws, r, 3)
                d_val = _cell(ws, r, 4)

                if a_val and str(a_val).strip() and b_val is None and c_val is None:
                    # Likely a brand header
                    candidate = str(a_val).strip()
                    skip_detail_kw = ["brands purchased", "purchase survey", "survey detail"]
                    if not any(kw in candidate.lower() for kw in skip_detail_kw):
                        if current_detail and current_brand:
                            details[current_brand] = current_detail
                        current_brand = candidate
                        current_detail = BrandPurchaseDetail(brand=candidate)
                    continue

                if current_detail is None:
                    continue

                if not a_val:
                    continue
                ll = str(a_val).lower()

                # Market/brand units
                if "market units" in ll:
                    current_detail.market_units = _float(b_val)
                elif "units purchased" in ll and "market" not in ll:
                    current_detail.brand_units = _float(b_val)

                # Segment data: look for "Cold", "Cough", "Allergy" labels
                # Demographic data: "Young Singles", etc.
                # These have 3 columns: % of Market, Brand Share, % of Brand
                seg_map = {
                    "cold": ("cold_market_pct", "cold_brand_share_pct", "cold_brand_pct"),
                    "cough": ("cough_market_pct", "cough_brand_share_pct", "cough_brand_pct"),
                    "allergy": ("allergy_market_pct", "allergy_brand_share_pct", "allergy_brand_pct"),
                    "young single": ("young_singles_market_pct", "young_singles_brand_share_pct", "young_singles_brand_pct"),
                    "young fam": ("young_families_market_pct", "young_families_brand_share_pct", "young_families_brand_pct"),
                    "mature": ("mature_families_market_pct", "mature_families_brand_share_pct", "mature_families_brand_pct"),
                    "empty": ("empty_nesters_market_pct", "empty_nesters_brand_share_pct", "empty_nesters_brand_pct"),
                    "retired": ("retired_market_pct", "retired_brand_share_pct", "retired_brand_pct"),
                }

                for key, (f1, f2, f3) in seg_map.items():
                    if key in ll:
                        setattr(current_detail, f1, _float(b_val))
                        setattr(current_detail, f2, _float(c_val))
                        setattr(current_detail, f3, _float(d_val))
                        break

            if current_detail and current_brand:
                details[current_brand] = current_detail

    wb.close()
    return summaries, details


# ===================================================================
# MAIN LOADER
# ===================================================================

def load_year(year: int, downloads_dir: Path | None = None) -> YearData:
    """Load and parse all available reports for a given simulation year.

    Args:
        year: Simulation year (0 = Start, 1 = Year1, etc.)
        downloads_dir: Override the default downloads directory.

    Returns:
        YearData with all parsed reports.
    """
    global DOWNLOADS_DIR
    original_dir = DOWNLOADS_DIR
    if downloads_dir:
        DOWNLOADS_DIR = downloads_dir

    try:
        yd = YearData(year=year)

        yd.performance_summary = parse_performance_summary(year)
        yd.income_statement = parse_income_statement(year)
        yd.product_contributions = parse_product_contribution(year)
        yd.sales_reports = parse_sales_report(year)
        yd.promotion_reports = parse_promotion_report(year)
        yd.dashboard = parse_dashboard(year)
        yd.brand_formulations = parse_brand_formulations(year)
        yd.symptoms_reported = parse_symptoms_reported(year)
        yd.industry_outlook = parse_industry_outlook(year)
        yd.advertising = parse_advertising(year)
        yd.brand_perceptions = parse_brand_perceptions(year)
        yd.brand_awareness = parse_brand_awareness(year)
        yd.shelf_space = parse_shelf_space(year)
        yd.promotions = parse_promotion(year)
        yd.conjoint_analysis = parse_conjoint_analysis(year)
        yd.satisfaction = parse_satisfaction(year)
        yd.sales_force = parse_sales_force(year)
        pricing, discount_details = parse_pricing(year)
        yd.pricing = pricing
        yd.channel_discount_details = discount_details
        yd.decision_criteria = parse_decision_criteria(year)
        yd.recommendations = parse_recommendations(year)
        yd.operating_statistics = parse_operating_statistics(year)
        yd.manufacturer_sales = parse_manufacturer_sales(year)
        yd.channel_sales = parse_channel_sales(year)
        yd.purchase_intentions = parse_purchase_intentions(year)
        yd.shopping_habits = parse_shopping_habits(year)
        yd.trade_offs = parse_trade_offs(year)
        bp_summary, bp_detail = parse_brands_purchased(year)
        yd.brands_purchased_summary = bp_summary
        yd.brands_purchased_detail = bp_detail

        return yd
    finally:
        DOWNLOADS_DIR = original_dir


def load_all_years(
    years: list[int] | None = None, downloads_dir: Path | None = None
) -> dict[int, YearData]:
    """Load data for multiple years.

    Args:
        years: List of years to load. Defaults to [0, 1].
        downloads_dir: Override the default downloads directory.

    Returns:
        Dict mapping year number to YearData.
    """
    if years is None:
        years = [0, 1]
    return {y: load_year(y, downloads_dir=downloads_dir) for y in years}


if __name__ == "__main__":
    import json
    from dataclasses import asdict

    print("Loading Year 0...")
    y0 = load_year(0)
    print("Loading Year 1...")
    y1 = load_year(1)

    def _summarize(yd: YearData):
        d = asdict(yd)
        for key, val in d.items():
            if val is None:
                print(f"  {key}: None (file not found)")
            elif isinstance(val, dict):
                print(f"  {key}: {len(val)} entries - keys: {list(val.keys())[:5]}")
            elif isinstance(val, list):
                print(f"  {key}: {len(val)} items")
            else:
                # Scalar or dataclass - count non-None fields
                if isinstance(val, dict):
                    non_none = sum(1 for v in val.values() if v is not None)
                    print(f"  {key}: {non_none} fields populated")
                else:
                    print(f"  {key}: {val}")

    print("\n=== Year 0 ===")
    _summarize(y0)
    print("\n=== Year 1 ===")
    _summarize(y1)

    # Spot-check key values
    print("\n=== Spot Checks ===")
    if y0.performance_summary:
        ps = y0.performance_summary
        print(f"Y0 Stock Price: {ps.stock_price}")
        print(f"Y0 Net Income: {ps.net_income}")
        print(f"Y0 Unit Sales: {ps.unit_sales}")
    if y0.income_statement:
        print(f"Y0 Mfr Sales: {y0.income_statement.manufacturer_sales}")
    if y0.brand_formulations:
        print(f"Y0 Brands: {list(y0.brand_formulations.keys())}")
    if y1.conjoint_analysis:
        print(f"Y1 Conjoint price points: {y1.conjoint_analysis.price_points}")
